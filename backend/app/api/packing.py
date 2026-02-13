"""Packing Proof API: batches, store jobs, allocation upload, photos."""
import csv
import io
import os
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session

from app.core.db import get_db
from app.core.config import settings
from app.api.permissions import require_admin, require_packer, get_current_user
from app.models.user import User
from app.models.job import Job
from app.models.packing_batch import PackingBatch
from app.models.packing_store_job import PackingStoreJob
from app.models.packing_store_line_item import PackingStoreLineItem
from app.models.file import File
from app.models.file_link import FileLink
from app.models.base import new_id
from app.schemas.packing import (
    PackingBatchCreate,
    PackingBatchOut,
    PackingStoreJobOut,
    PackingStoreJobDetailOut,
    PackingStoreJobUpdate,
    PackingStoreLineItemOut,
)

router = APIRouter(prefix="/packing", tags=["packing"])

# Uploads dir relative to backend root (backend = parent of app)
_BACKEND_ROOT = Path(__file__).resolve().parent.parent.parent
UPLOADS_DIR = _BACKEND_ROOT / settings.UPLOADS_DIR


def _ensure_uploads():
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    (UPLOADS_DIR / "packing").mkdir(exist_ok=True)


# ---- Batches (admin) ----
@router.get("/batches", response_model=list[PackingBatchOut])
def list_batches(
    job_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    q = db.query(PackingBatch).order_by(PackingBatch.created_at.desc())
    if job_id:
        q = q.filter(PackingBatch.job_id == job_id)
    return q.all()


@router.post("/batches", response_model=PackingBatchOut)
def create_batch(
    payload: PackingBatchCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    job = db.query(Job).filter(Job.id == payload.job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    batch = PackingBatch(
        id=new_id(),
        job_id=payload.job_id,
        name=payload.name or f"Batch {payload.job_id[:8]}",
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return batch


def _parse_allocation_rows(content: bytes, filename: str) -> list[dict]:
    """Parse CSV or xlsx into rows of store_name, component, description, qty. Suppress qty 0."""
    rows: list[dict] = []
    lower = filename.lower() if filename else ""
    if lower.endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            # Normalize keys (strip, lowercase)
            row = {k.strip().lower().replace(" ", "_"): v.strip() if isinstance(v, str) else v for k, v in r.items()}
            store = row.get("store_name") or row.get("store") or row.get("storename") or ""
            component = row.get("component") or row.get("component_name") or ""
            description = row.get("description") or row.get("desc") or ""
            try:
                qty = int(float(row.get("qty") or row.get("quantity") or 0))
            except (ValueError, TypeError):
                qty = 0
            if qty <= 0:
                continue
            if not store:
                continue
            rows.append({"store_name": store, "component": component, "description": description, "qty": qty})
        return rows
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=400, detail="xlsx support requires openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2):
            vals = [c.value for c in row]
            row_dict = dict(zip(headers, (v if v is None else str(v).strip() for v in vals)))
            store = row_dict.get("store_name") or row_dict.get("store") or row_dict.get("storename") or ""
            component = row_dict.get("component") or row_dict.get("component_name") or ""
            description = row_dict.get("description") or row_dict.get("desc") or ""
            try:
                qty = int(float(row_dict.get("qty") or row_dict.get("quantity") or 0))
            except (ValueError, TypeError):
                qty = 0
            if qty <= 0:
                continue
            if not store:
                continue
            rows.append({"store_name": store, "component": component, "description": description, "qty": qty})
        return rows
    raise HTTPException(status_code=400, detail="Upload a .csv or .xlsx file")


@router.post("/batches/{batch_id}/upload-allocation")
def upload_allocation(
    batch_id: str,
    file: UploadFile = File(),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    batch = db.query(PackingBatch).filter(PackingBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file")
    content = file.file.read()
    rows = _parse_allocation_rows(content, file.filename)
    # Group by store_name -> one PackingStoreJob per store
    from collections import defaultdict
    by_store: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_store[r["store_name"]].append(r)
    created = 0
    for store_name, line_rows in by_store.items():
        store_job = PackingStoreJob(
            id=new_id(),
            batch_id=batch_id,
            store_name=store_name,
            status="pending",
        )
        db.add(store_job)
        db.flush()
        for r in line_rows:
            line = PackingStoreLineItem(
                id=new_id(),
                store_job_id=store_job.id,
                component=r.get("component", ""),
                description=r.get("description", ""),
                qty=r.get("qty", 0),
            )
            db.add(line)
            created += 1
    db.commit()
    return {"batch_id": batch_id, "store_jobs_created": len(by_store), "line_items_created": created}


# ---- Store jobs (mobile / packer) ----
@router.get("/store-jobs", response_model=list[PackingStoreJobOut])
def list_store_jobs(
    batch_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_packer),
):
    q = db.query(PackingStoreJob).order_by(PackingStoreJob.store_name)
    if batch_id:
        q = q.filter(PackingStoreJob.batch_id == batch_id)
    if status:
        q = q.filter(PackingStoreJob.status == status)
    return q.all()


@router.get("/store-jobs/{store_job_id}", response_model=PackingStoreJobDetailOut)
def get_store_job(
    store_job_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_packer),
):
    job = db.query(PackingStoreJob).filter(PackingStoreJob.id == store_job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Store job not found")
    items = db.query(PackingStoreLineItem).filter(PackingStoreLineItem.store_job_id == store_job_id).all()
    out = PackingStoreJobDetailOut.model_validate(job)
    out.line_items = [PackingStoreLineItemOut.model_validate(i) for i in items]
    return out


@router.patch("/store-jobs/{store_job_id}", response_model=PackingStoreJobOut)
def update_store_job(
    store_job_id: str,
    payload: PackingStoreJobUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_packer),
):
    job = db.query(PackingStoreJob).filter(PackingStoreJob.id == store_job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Store job not found")
    if payload.status is not None:
        job.status = payload.status
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc)
        if payload.status == "packed" and job.packed_at is None:
            job.packed_at = now
        if payload.status == "dispatched" and job.dispatched_at is None:
            job.dispatched_at = now
    if payload.box_count is not None:
        job.box_count = payload.box_count
    if payload.notes is not None:
        job.notes = payload.notes
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def _mime_from_filename(filename: str) -> str:
    ext = (filename or "").split(".")[-1].lower()
    mimes = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "gif": "image/gif", "webp": "image/webp"}
    return mimes.get(ext, "application/octet-stream")


@router.post("/store-jobs/{store_job_id}/photos")
def upload_store_job_photos(
    store_job_id: str,
    files: list[UploadFile] = File(),
    db: Session = Depends(get_db),
    user: User = Depends(require_packer),
):
    job = db.query(PackingStoreJob).filter(PackingStoreJob.id == store_job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Store job not found")
    _ensure_uploads()
    uploaded = []
    for f in files or []:
        if not f.filename:
            continue
        content = f.file.read()
        if not content:
            continue
        file_id = new_id()
        ext = (f.filename or "").split(".")[-1].lower() or "bin"
        if ext not in ("jpg", "jpeg", "png", "gif", "webp"):
            ext = "bin"
        storage_key = f"packing/{file_id}.{ext}"
        path = UPLOADS_DIR / storage_key
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(content)
        mime = _mime_from_filename(f.filename or "")
        file_row = File(
            id=file_id,
            storage_key=storage_key,
            mime=mime,
            size=len(content),
            sha256="",  # optional: hashlib.sha256(content).hexdigest()
            uploaded_by=user.id,
        )
        db.add(file_row)
        db.flush()
        link = FileLink(
            id=new_id(),
            file_id=file_id,
            entity_type="packing_store_job",
            entity_id=store_job_id,
            tag="packed_photo",
        )
        db.add(link)
        uploaded.append({"file_id": file_id, "storage_key": storage_key})
    db.commit()
    return {"store_job_id": store_job_id, "uploaded": uploaded}
